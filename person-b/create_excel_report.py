"""
Create Excel report for Oracle behavior analysis
Uses REAL Chainlink data from chainlink_history.csv
"""

import csv
import os
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def load_chainlink_data(filename="chainlink_history.csv"):
    """Load Chainlink oracle data from CSV"""
    filepath = os.path.join(SCRIPT_DIR, filename)
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append({
                'roundId': int(row['roundId']),
                'timestamp': int(row['timestamp']),
                'price': float(row['price'])
            })
    return data


def calculate_price_changes(chainlink_data):
    """Calculate price changes between consecutive Oracle updates"""
    price_changes = []

    for i in range(1, len(chainlink_data)):
        prev = chainlink_data[i - 1]
        curr = chainlink_data[i]

        price_change_pct = (curr['price'] - prev['price']) / prev['price'] * 100
        abs_change = abs(price_change_pct)
        time_delta = curr['timestamp'] - prev['timestamp']

        dt = datetime.fromtimestamp(curr['timestamp'], tz=timezone.utc)

        price_changes.append({
            'datetime': dt,
            'date': dt.strftime('%Y-%m-%d'),
            'time': dt.strftime('%H:%M:%S'),
            'hour': dt.hour,
            'prev_price': prev['price'],
            'curr_price': curr['price'],
            'change_pct': price_change_pct,
            'abs_change': abs_change,
            'direction': 'UP' if price_change_pct > 0 else 'DOWN',
            'time_delta_sec': time_delta
        })

    return price_changes


def create_excel_report(price_changes, output_file="docs/oracle_behavior.xlsx"):
    """Create Excel report with multiple sheets"""

    filepath = os.path.join(SCRIPT_DIR, output_file)
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== Sheet 1: Übersicht ==========
    ws1 = wb.active
    ws1.title = "Übersicht"

    # Calculate statistics
    threshold = 0.35
    significant = [e for e in price_changes if e['abs_change'] >= threshold]
    prices = [e['curr_price'] for e in price_changes]
    abs_changes = [e['abs_change'] for e in significant]

    timestamps = [e['datetime'].timestamp() for e in price_changes]
    days = (max(timestamps) - min(timestamps)) / 86400

    up_moves = len([e for e in significant if e['direction'] == 'UP'])
    down_moves = len([e for e in significant if e['direction'] == 'DOWN'])

    # Overview data
    overview_data = [
        ["Metrik", "Wert", "Beschreibung"],
        ["Analysezeitraum", f"{days:.1f} Tage", "Zeitraum der Datenerfassung"],
        ["Total Oracle Updates", len(price_changes), "Anzahl der Oracle-Preisänderungen"],
        ["Signifikante Moves (>=0.35%)", len(significant), "Trading-relevante Ereignisse"],
        ["Signal Rate", f"{len(significant)/len(price_changes)*100:.2f}%", "Anteil signifikanter Moves"],
        ["Trading Signale/Tag", f"{len(significant)/days:.1f}", "Durchschnittliche Signale pro Tag"],
        ["Mean Price Change", f"{sum(abs_changes)/len(abs_changes):.3f}%", "Durchschnittliche Preisänderung"],
        ["Max Price Change", f"{max(abs_changes):.3f}%", "Größte Preisänderung"],
        ["Min Price (Zeitraum)", f"${min(prices):,.2f}", "Niedrigster BTC-Preis"],
        ["Max Price (Zeitraum)", f"${max(prices):,.2f}", "Höchster BTC-Preis"],
        ["Upward Signals", up_moves, "Anzahl aufwärts-gerichteter Signale"],
        ["Downward Signals", down_moves, "Anzahl abwärts-gerichteter Signale"],
        ["Up/Down Ratio", f"{up_moves/down_moves:.2f}" if down_moves > 0 else "N/A", "Verhältnis Up zu Down"],
    ]

    for row_idx, row_data in enumerate(overview_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

    # Adjust column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 40

    # ========== Sheet 2: Top Events ==========
    ws2 = wb.create_sheet("Top Events")

    # Sort by absolute change and get top 20
    top_events = sorted(price_changes, key=lambda x: x['abs_change'], reverse=True)[:20]

    headers = ["Rank", "Datum", "Uhrzeit (UTC)", "Preis vorher ($)", "Preis nachher ($)", "Change (%)", "Richtung"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, event in enumerate(top_events, 2):
        ws2.cell(row=row_idx, column=1, value=row_idx-1).border = border
        ws2.cell(row=row_idx, column=2, value=event['date']).border = border
        ws2.cell(row=row_idx, column=3, value=event['time']).border = border
        ws2.cell(row=row_idx, column=4, value=f"${event['prev_price']:,.2f}").border = border
        ws2.cell(row=row_idx, column=5, value=f"${event['curr_price']:,.2f}").border = border

        change_cell = ws2.cell(row=row_idx, column=6, value=f"{event['change_pct']:+.3f}%")
        change_cell.border = border
        if event['change_pct'] > 0:
            change_cell.font = Font(color="008000")  # Green
        else:
            change_cell.font = Font(color="FF0000")  # Red

        dir_cell = ws2.cell(row=row_idx, column=7, value=event['direction'])
        dir_cell.border = border
        if event['direction'] == 'UP':
            dir_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            dir_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Adjust column widths
    for col_idx, width in enumerate([8, 12, 14, 18, 18, 14, 12], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ========== Sheet 3: Peak Hours ==========
    ws3 = wb.create_sheet("Peak Hours")

    # Count events by hour
    hourly_counts = {}
    hourly_up = {}
    hourly_down = {}
    for h in range(24):
        hourly_counts[h] = 0
        hourly_up[h] = 0
        hourly_down[h] = 0

    for event in significant:
        h = event['hour']
        hourly_counts[h] += 1
        if event['direction'] == 'UP':
            hourly_up[h] += 1
        else:
            hourly_down[h] += 1

    headers = ["Stunde (UTC)", "Total Events", "Up", "Down", "% des Tages", "Visualisierung"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    total_events = sum(hourly_counts.values())
    max_events = max(hourly_counts.values())

    for row_idx, hour in enumerate(range(24), 2):
        count = hourly_counts[hour]
        pct = count / total_events * 100 if total_events > 0 else 0
        bar = "#" * int(count / max_events * 20) if max_events > 0 else ""

        ws3.cell(row=row_idx, column=1, value=f"{hour:02d}:00").border = border
        ws3.cell(row=row_idx, column=2, value=count).border = border
        ws3.cell(row=row_idx, column=3, value=hourly_up[hour]).border = border
        ws3.cell(row=row_idx, column=4, value=hourly_down[hour]).border = border
        ws3.cell(row=row_idx, column=5, value=f"{pct:.1f}%").border = border
        ws3.cell(row=row_idx, column=6, value=bar).border = border

        # Highlight peak hours
        if count >= max_events * 0.8:
            for col in range(1, 7):
                ws3.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )

    # Adjust column widths
    for col_idx, width in enumerate([14, 14, 8, 8, 14, 25], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    # Save
    wb.save(filepath)
    print(f"[OK] Excel report saved: {filepath}")

    return filepath


if __name__ == "__main__":
    print("Loading Chainlink data...")
    data = load_chainlink_data()
    print(f"Loaded {len(data)} records")

    print("Calculating price changes...")
    changes = calculate_price_changes(data)
    print(f"Calculated {len(changes)} price changes")

    print("Creating Excel report...")
    create_excel_report(changes)
    print("Done!")
